from pathlib import Path
from datetime import datetime
import os
import shutil
from getpass import getuser
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from utils import YN

# ------------------------------------------------------------
# Configuration
# ------------------------------------------------------------

CODE_DIRECTORY = Path(__file__).resolve().parent
AUTOMATION_DIRECTORY = CODE_DIRECTORY.parent
DOCUMENTS_DIRECTORY = AUTOMATION_DIRECTORY.parent
REPORTS_DIRECTORY = AUTOMATION_DIRECTORY / "Reports"

# Top-level folder containing the PDFs.
TARGET_DIRECTORY = DOCUMENTS_DIRECTORY / "Target"

# Folders used for rejected and held documents.
REJECTED_DIRECTORY = TARGET_DIRECTORY / "Rejected"
HOLD_DIRECTORY = TARGET_DIRECTORY / "Hold"

'''------------------------------------------------------------'''

# Set this to None to select the newest timestamped report.
#
# To select a specific report, replace None with:
#
# RESULTS_WORKBOOK_OVERRIDE = (
#     REPORTS_DIRECTORY
#     / "RenameResults_2026-07-27_15-34-18.xlsx"
# )

RESULTS_WORKBOOK_OVERRIDE = None        ;'''Edit this to select a specific report, or leave as "None" to select the newest timestamped report.'''
DRY_RUN = YN("Would you like to perform a dry run? (y/n): ")                          ;'''Set to True to simulate approval processing without renaming or moving files. Set to False to perform real file operations.'''

'''------------------------------------------------------------'''
CURRENT_USER = getuser()

REPORT_NAME_PATTERN = re.compile(
    r"^RenameResults_"
    r"(\d{4}-\d{2}-\d{2}_\d{2}-\d{2}-\d{2})"
    r"\.xlsx$",
    re.IGNORECASE
)


def select_results_workbook():
    """
    Return the explicitly configured workbook, or the report with
    the latest timestamp embedded in its filename.
    """

    if RESULTS_WORKBOOK_OVERRIDE is not None:
        selected_path = Path(
            RESULTS_WORKBOOK_OVERRIDE
        )

        return (
            selected_path,
            "Explicit workbook override"
        )

    candidates = []

    for workbook_path in REPORTS_DIRECTORY.glob(
        "RenameResults_*.xlsx"
    ):
        if workbook_path.name.startswith("~$"):
            continue

        match = REPORT_NAME_PATTERN.fullmatch(
            workbook_path.name
        )

        if not match:
            continue

        try:
            report_timestamp = datetime.strptime(
                match.group(1),
                "%Y-%m-%d_%H-%M-%S"
            )

        except ValueError:
            continue

        candidates.append(
            (
                report_timestamp,
                workbook_path
            )
        )

    if not candidates:
        raise FileNotFoundError(
            "No timestamped results workbooks were found in:\n"
            f"{REPORTS_DIRECTORY}"
        )

    candidates.sort(
        key=lambda item: item[0],
        reverse=True
    )

    return (
        candidates[0][1],
        "Automatically selected newest timestamped report"
    )


RESULTS_WORKBOOK, WORKBOOK_SELECTION_REASON = (
    select_results_workbook()
)


# ------------------------------------------------------------
# Display configuration
# ------------------------------------------------------------

print("Approval locations:")
print(f"  Target directory:  {TARGET_DIRECTORY}")
print(f"  Rejected folder:   {REJECTED_DIRECTORY}")
print(f"  Hold folder:       {HOLD_DIRECTORY}")
print(f"  Dry run:           {DRY_RUN}")
print(f"  Selected because:  {WORKBOOK_SELECTION_REASON}")
print(f"  Current user:      {CURRENT_USER}")
print(f"  Dry run:           {DRY_RUN}")
print(f"  Selected results workbook:  {RESULTS_WORKBOOK}")

print()

check = YN("Are you happy to continue with these settings? (y/n): ")
if not check:
    print("Exiting.")
    exit(0)

# ------------------------------------------------------------
# Validate paths
# ------------------------------------------------------------

if not TARGET_DIRECTORY.exists():
    raise FileNotFoundError(
        "The target directory could not be found:\n"
        f"{TARGET_DIRECTORY}"
    )

if not TARGET_DIRECTORY.is_dir():
    raise NotADirectoryError(
        "The configured target path is not a directory:\n"
        f"{TARGET_DIRECTORY}"
    )

if not RESULTS_WORKBOOK.exists():
    raise FileNotFoundError(
        "The reviewed results workbook could not be found:\n"
        f"{RESULTS_WORKBOOK}"
    )

if not DRY_RUN:
    try:
        with open(
            RESULTS_WORKBOOK,
            "r+b"
        ):
            pass

    except PermissionError as error:
        raise PermissionError(
            "The results workbook cannot be updated. "
            "It may currently be open in Excel:\n"
            f"{RESULTS_WORKBOOK}"
        ) from error


# ------------------------------------------------------------
# Load reviewed results
# ------------------------------------------------------------

results = pd.read_excel(
    RESULTS_WORKBOOK,
    sheet_name="All Results",
    engine="openpyxl"
)


required_columns = [
    "Original File",
    "Generated Filename",
    "Proposed Path",
    "Status",
    "Manual Decision",
    "Approved Filename",
    "Approval Result",
    "Approved By",
    "Approved At",
    "Final Path"
]


missing_columns = [
    column
    for column in required_columns
    if column not in results.columns
]


if missing_columns:
    raise ValueError(
        "The workbook is missing required columns: "
        + ", ".join(missing_columns)
    )

approval_workbook = load_workbook(
    RESULTS_WORKBOOK
)

results_sheet = approval_workbook[
    "All Results"
]

if "Audit Log" not in approval_workbook.sheetnames:
    raise ValueError(
        "The selected workbook does not contain an Audit Log sheet."
    )

audit_sheet = approval_workbook[
    "Audit Log"
]

result_headers = {
    cell.value: cell.column
    for cell in results_sheet[1]
}

# ------------------------------------------------------------
# Clean manual decisions
# ------------------------------------------------------------

results["Manual Decision"] = (
    results["Manual Decision"]
    .fillna("")
    .astype(str)
    .str.strip()
    .str.upper()
)


valid_decisions = {
    "",
    "APPROVE",
    "REJECT",
    "HOLD"
}


invalid_decision_rows = results[
    ~results["Manual Decision"].isin(valid_decisions)
]


if not invalid_decision_rows.empty:

    print(
        "WARNING: Some rows contain invalid manual decisions."
    )

    print(
        "Only APPROVE, REJECT, HOLD, or blank are supported."
    )

    for _, row in invalid_decision_rows.iterrows():
        print(
            f"  {row['Original File']}: "
            f"{row['Manual Decision']}"
        )

    print()


# ------------------------------------------------------------
# Decision summary
# ------------------------------------------------------------

decision_counts = (
    results["Manual Decision"]
    .replace("", "BLANK")
    .value_counts(dropna=False)
)


print("Manual decision summary:")

for decision, count in decision_counts.items():
    print(f"  {decision}: {count}")

print()


# ------------------------------------------------------------
# Helper functions
# ------------------------------------------------------------

def paths_are_equal(first_path, second_path):
    """
    Compare two paths in a Windows-safe, case-insensitive manner.
    """

    return (
        os.path.normcase(
            str(first_path.resolve())
        )
        ==
        os.path.normcase(
            str(second_path.resolve())
        )
    )


def source_is_inside_target(source_path):
    """
    Return True only when the source file is inside the configured
    Target directory.
    """

    try:
        source_path.resolve().relative_to(
            TARGET_DIRECTORY.resolve()
        )

        return True

    except ValueError:
        return False

def source_is_inside_directory(
    source_path,
    directory_path
):
    """
    Return True when source_path is located inside directory_path,
    including any of its subfolders.
    """

    try:
        source_path.resolve().relative_to(
            directory_path.resolve()
        )

        return True

    except ValueError:
        return False
    


def validate_source(source_path):
    """
    Check whether a source path is eligible for processing.

    Returns:
        tuple: (is_valid, message)
    """

    if not source_path.exists():
        return (
            False,
            "Source file does not exist"
        )

    if not source_path.is_file():
        return (
            False,
            "Source path is not a file"
        )

    if source_path.suffix.lower() != ".pdf":
        return (
            False,
            "Source is not a PDF"
        )

    if not source_is_inside_target(source_path):
        return (
            False,
            "Source file is outside the configured Target directory"
        )

    return (
        True,
        ""
    )


def move_review_file(
    source_path,
    destination_directory,
    decision_name
):
    """
    Move a rejected or held PDF into the corresponding folder.

    Returns:
        tuple: (success, message, resulting_path)
    """

    destination_path = (
        destination_directory
        / source_path.name
    )

    print()
    print(f"{decision_name} source: {source_path}")
    print(f"{decision_name} target: {destination_path}")

    is_valid, validation_message = validate_source(
        source_path
    )

    if not is_valid:
        print(f"SKIPPED: {validation_message}")

        return (
            False,
            validation_message,
            None
        )

    if paths_are_equal(
        source_path,
        destination_path
    ):
        message = (
            f"File is already in the {decision_name} folder"
        )

        print(f"SKIPPED: {message}")

        return (
            True,
            message,
            destination_path
        )

    if destination_path.exists():
        message = (
            f"Destination already exists: {destination_path}"
        )

        print(f"SKIPPED: {message}")

        return (
            False,
            message,
            None
        )

    if DRY_RUN:
        message = (
            f"Dry run - would move file to "
            f"{decision_name} folder"
        )

        print(message)

        return (
            True,
            message,
            destination_path
        )

    destination_directory.mkdir(
        parents=True,
        exist_ok=True
    )

    shutil.move(
        str(source_path),
        str(destination_path)
    )

    message = f"Moved to {decision_name} folder"

    print(message)

    return (
        True,
        message,
        destination_path
    )


def rename_approved_file(
    source_path,
    generated_filename,
    approved_filename
):
    """
    Rename a manually approved PDF.

    Approved Filename takes priority over Generated Filename.

    If the source is currently inside Hold, the approved file is
    moved back into the main Target directory while being renamed.

    Files approved from elsewhere are renamed in their current
    directory.

    Returns:
        tuple: (success, message, resulting_path)
    """

    if (
        pd.notna(approved_filename)
        and str(approved_filename).strip()
    ):
        final_filename = str(
            approved_filename
        ).strip()

        filename_source = "Approved Filename"

    elif (
        pd.notna(generated_filename)
        and str(generated_filename).strip()
    ):
        final_filename = str(
            generated_filename
        ).strip()

        filename_source = "Generated Filename"

    else:
        message = "No filename supplied"

        print()
        print(f"Approved source: {source_path}")
        print(f"SKIPPED: {message}")

        return (
            False,
            message,
            None
        )

    # Remove the extension if the reviewer included it.
    if final_filename.lower().endswith(".pdf"):
        final_filename = final_filename[:-4]

    # Prevent an accidental path from being entered as a filename.
    final_filename = Path(
        final_filename
    ).name

    # Files approved from Hold should return to the main designated
    # target directory.
    if source_is_inside_directory(
        source_path,
        HOLD_DIRECTORY
    ):
        destination_directory = (
            TARGET_DIRECTORY
        )

        destination_reason = (
            "Approved file will be moved out of Hold"
        )

        success_message = (
            "Approved file renamed and moved out of Hold"
        )

    else:
        destination_directory = (
            source_path.parent
        )

        destination_reason = (
            "Approved file will remain in its current folder"
        )

        success_message = (
            "Renamed approved file"
        )

    destination_path = (
        destination_directory
        / f"{final_filename}.pdf"
    )

    print()
    print(f"Approved source: {source_path}")
    print(f"Approved target: {destination_path}")
    print(f"Filename source: {filename_source}")
    print(f"Destination rule: {destination_reason}")

    is_valid, validation_message = validate_source(
        source_path
    )

    if not is_valid:
        print(f"SKIPPED: {validation_message}")

        return (
            False,
            validation_message,
            None
        )

    if paths_are_equal(
        source_path,
        destination_path
    ):
        message = "File is already correctly named"

        print(f"SKIPPED: {message}")

        return (
            True,
            message,
            destination_path
        )

    if destination_path.exists():
        message = (
            f"Destination already exists: {destination_path}"
        )

        print(f"SKIPPED: {message}")

        return (
            False,
            message,
            None
        )

    if DRY_RUN:
        if source_is_inside_directory(
            source_path,
            HOLD_DIRECTORY
        ):
            message = (
                "Dry run - would rename approved file "
                "and move it out of Hold"
            )

        else:
            message = (
                "Dry run - would rename approved file"
            )

        print(message)

        return (
            True,
            message,
            destination_path
        )

    destination_directory.mkdir(
        parents=True,
        exist_ok=True
    )

    shutil.move(
        str(source_path),
        str(destination_path)
    )

    print(success_message)

    return (
        True,
        success_message,
        destination_path
    )

def get_current_source_path(row):
    """
    Return the latest known location of a PDF.

    Search order:
    1. Final Path from a previous approval operation.
    2. Original File when it still exists.
    3. Proposed Path only when the workbook indicates that a real
       automatic rename was completed successfully.
    4. Original File as the final fallback for a useful validation
       error.
    """

    final_path_value = row.get(
        "Final Path"
    )

    if (
        pd.notna(final_path_value)
        and str(final_path_value).strip()
    ):
        final_path = Path(
            str(final_path_value).strip()
        )

        if final_path.exists():
            return final_path

    original_path = Path(
        str(row["Original File"])
    )

    if original_path.exists():
        return original_path

    proposed_path_value = row.get(
        "Proposed Path"
    )

    rename_result = normalise_cell_text(
        row.get("Rename Result")
    )

    status = normalise_cell_text(
        row.get("Status")
    ).upper()

    automatic_rename_completed = (
        status == "SUCCESS"
        and rename_result
        and not rename_result.lower().startswith(
            "dry run"
        )
        and not rename_result.lower().startswith(
            "not renamed"
        )
    )

    if (
        automatic_rename_completed
        and pd.notna(proposed_path_value)
        and str(proposed_path_value).strip()
    ):
        proposed_path = Path(
            str(proposed_path_value).strip()
        )

        if proposed_path.exists():
            return proposed_path

    return original_path

def normalise_cell_text(value):
    """
    Return a trimmed string for a possibly blank Excel value.
    """

    if pd.isna(value):
        return ""

    return str(value).strip()


def approval_is_permanently_completed(row):
    """
    Completed approvals and rejections must not be repeated.

    HOLD and FAILED results remain eligible for later reconsideration.
    """

    approval_result = normalise_cell_text(
        row.get("Approval Result")
    ).upper()

    return approval_result.startswith(
        "COMPLETED -"
    )


def write_approval_result(
    excel_row,
    result_text,
    processed_at,
    final_path=None
):
    """
    Write one real approval result into All Results.
    """

    values = {
        "Approval Result": result_text,
        "Approved By": CURRENT_USER,
        "Approved At": processed_at.strftime(
            "%Y-%m-%d %H:%M:%S"
        ),
        "Final Path": (
            str(Path(final_path).resolve())
            if final_path is not None
            else ""
        )
    }

    for heading, value in values.items():
        column_number = result_headers.get(
            heading
        )

        if column_number:
            results_sheet.cell(
                row=excel_row,
                column=column_number
            ).value = value


def append_audit_event(
    timestamp,
    details
):
    """
    Append an approval-run event to the Audit Log sheet.
    """

    audit_sheet.append(
        [
            timestamp.strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
            "APPROVAL_RUN",
            CURRENT_USER,
            False,
            "Rename_Approval.py",
            details
        ]
    )

    audit_table = audit_sheet.tables.get(
        "tblAuditLog"
    )

    if audit_table is not None:
        audit_table.ref = (
            f"A1:"
            f"{get_column_letter(audit_sheet.max_column)}"
            f"{audit_sheet.max_row}"
        )


# ------------------------------------------------------------
# Process manual decisions
# ------------------------------------------------------------

approved_count = 0
rejected_count = 0
hold_count = 0
blank_count = 0
invalid_count = 0
successful_operations = 0
skipped_operations = 0


for excel_row, (_, row) in enumerate(
    results.iterrows(),
    start=2
):
    decision = row["Manual Decision"]

    source_path = get_current_source_path(
        row
    )

    generated_filename = row.get(
        "Generated Filename"
    )

    approved_filename = row.get(
        "Approved Filename"
    )

    # Completed approvals and rejections are final and must not
    # be applied again.
    if approval_is_permanently_completed(row):
        skipped_operations += 1

        print()
        print(f"Previously completed: {source_path}")
        print("SKIPPED: Approval action already completed")

        continue

    # --------------------------------------------------------
    # APPROVE
    # --------------------------------------------------------

    if decision == "APPROVE":

        approved_count += 1

        success, message, resulting_path = (
            rename_approved_file(
                source_path=source_path,
                generated_filename=generated_filename,
                approved_filename=approved_filename
            )
        )

        if success:
            successful_operations += 1

            if not DRY_RUN:
                write_approval_result(
                    excel_row=excel_row,
                    result_text=(
                        f"COMPLETED - {message}"
                    ),
                    processed_at=datetime.now(),
                    final_path=resulting_path
                )

        else:
            skipped_operations += 1

            if not DRY_RUN:
                write_approval_result(
                    excel_row=excel_row,
                    result_text=(
                        f"FAILED - {message}"
                    ),
                    processed_at=datetime.now(),
                    final_path=None
                )

    # --------------------------------------------------------
    # REJECT
    # --------------------------------------------------------

    elif decision == "REJECT":

        rejected_count += 1

        success, message, resulting_path = (
            move_review_file(
                source_path=source_path,
                destination_directory=REJECTED_DIRECTORY,
                decision_name="Rejected"
            )
        )

        if success:
            successful_operations += 1

            if not DRY_RUN:
                write_approval_result(
                    excel_row=excel_row,
                    result_text=(
                        f"COMPLETED - {message}"
                    ),
                    processed_at=datetime.now(),
                    final_path=resulting_path
                )

        else:
            skipped_operations += 1

            if not DRY_RUN:
                write_approval_result(
                    excel_row=excel_row,
                    result_text=(
                        f"FAILED - {message}"
                    ),
                    processed_at=datetime.now(),
                    final_path=None
                )

    # --------------------------------------------------------
    # HOLD
    # --------------------------------------------------------

    elif decision == "HOLD":

        hold_count += 1

        success, message, resulting_path = (
            move_review_file(
                source_path=source_path,
                destination_directory=HOLD_DIRECTORY,
                decision_name="Hold"
            )
        )

        if success:
            successful_operations += 1

            if not DRY_RUN:
                write_approval_result(
                    excel_row=excel_row,
                    result_text=(
                        f"HOLD - {message}"
                    ),
                    processed_at=datetime.now(),
                    final_path=resulting_path
                )

        else:
            skipped_operations += 1

            if not DRY_RUN:
                write_approval_result(
                    excel_row=excel_row,
                    result_text=(
                        f"FAILED - {message}"
                    ),
                    processed_at=datetime.now(),
                    final_path=None
                )

    # --------------------------------------------------------
    # Blank
    # --------------------------------------------------------

    elif decision == "":

        blank_count += 1

        # Blank records remain available for future review.
        # No filesystem or workbook change is made.

    # --------------------------------------------------------
    # Invalid value
    # --------------------------------------------------------

    else:

        invalid_count += 1

        message = (
            f"Invalid manual decision: {decision}"
        )

        print()
        print(f"Invalid decision for: {source_path}")
        print(f"SKIPPED: {decision}")

        if not DRY_RUN:
            write_approval_result(
                excel_row=excel_row,
                result_text=(
                    f"FAILED - {message}"
                ),
                processed_at=datetime.now(),
                final_path=None
            )

# ------------------------------------------------------------
# Save real approval outcomes and append the run audit
# ------------------------------------------------------------

if not DRY_RUN:

    approval_finished_at = datetime.now()

    approval_details = (
        f"Approved={approved_count}; "
        f"Rejected={rejected_count}; "
        f"Hold={hold_count}; "
        f"Blank={blank_count}; "
        f"Invalid={invalid_count}; "
        f"Successful={successful_operations}; "
        f"Skipped={skipped_operations}"
    )

    append_audit_event(
        timestamp=approval_finished_at,
        details=approval_details
    )

    try:
        approval_workbook.save(
            RESULTS_WORKBOOK
        )

    except PermissionError as error:
        raise PermissionError(
            "File operations may have completed, but the workbook "
            "could not be updated. Ensure the workbook is closed "
            "in Excel before running real approval processing."
        ) from error

    print()
    print("Approval outcomes written to:")
    print(f"  {RESULTS_WORKBOOK.resolve()}")


# ------------------------------------------------------------
# Final summary
# ------------------------------------------------------------

print()
print("=" * 70)
print("Approval processing summary")
print("=" * 70)

print(f"Approved rows:       {approved_count}")
print(f"Rejected rows:       {rejected_count}")
print(f"Hold rows:           {hold_count}")
print(f"Blank decisions:     {blank_count}")
print(f"Invalid decisions:   {invalid_count}")
print(f"Successful actions:  {successful_operations}")
print(f"Skipped actions:     {skipped_operations}")
print(f"Dry run:             {DRY_RUN}")

if DRY_RUN:
    print()
    print(
        "No files were renamed or moved because DRY_RUN is enabled."
    )
else:
    print()
    print(
        "Approved files were renamed and rejected/held files "
        "were moved where possible."
    )