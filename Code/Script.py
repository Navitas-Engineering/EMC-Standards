import os
import pandas as pd
from datetime import datetime
from pathlib import Path
from getpass import getuser

from extraction import get_file_names
from processing import process_file, test_single_file
from extraction_continued import rename_file


from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------
# Configuration
# ------------------------------------------------------------

DRY_RUN = True
EXPORT_RESULTS = True

#Timestamp for workbook filename is generated automatically, so that multiple runs do not overwrite each other.
# Store the actual datetime object so callers can use .strftime() safely.
RUN_STARTED_AT = datetime.now()
CURRENT_USER = getuser()

# C:\Users\JoshuaDickens\Documents\Automation\Code
CODE_DIRECTORY = Path(__file__).resolve().parent

# C:\Users\JoshuaDickens\Documents\Automation
AUTOMATION_DIRECTORY = CODE_DIRECTORY.parent

REPORTS_DIRECTORY = (
    AUTOMATION_DIRECTORY
    / "Reports"
)

# C:\Users\JoshuaDickens\Documents
DOCUMENTS_DIRECTORY = AUTOMATION_DIRECTORY.parent

# C:\Users\JoshuaDickens\Documents\Target
TARGET_DIRECTORY = DOCUMENTS_DIRECTORY / "Target" #Enter name of the folder containing the PDFs to be renamed here. 

REJECTED_DIRECTORY = (
    TARGET_DIRECTORY
    / "Rejected"
)

# Store the report inside Automation:
# C:\Users\JoshuaDickens\Documents\Automation\RenameResults.xlsx
RESULTS_WORKBOOK = (
    REPORTS_DIRECTORY
    / (
        "RenameResults_"
        f"{RUN_STARTED_AT.strftime('%Y-%m-%d_%H-%M-%S')}"
        ".xlsx"
    )
)

print("Project locations:")
print(f"  Code directory:   {CODE_DIRECTORY}")
print(f"  Automation root:  {AUTOMATION_DIRECTORY}")
print(f"  Target directory: {TARGET_DIRECTORY}")
print(f"  Results workbook: {RESULTS_WORKBOOK}")
print(f"  Rejected folder:   {REJECTED_DIRECTORY}")
print(f"  Run started:       {RUN_STARTED_AT}")
print(f"  Created by:        {CURRENT_USER}")
print()

# ------------------------------------------------------------
# Processing
# ------------------------------------------------------------

if not TARGET_DIRECTORY.exists():
    raise FileNotFoundError(
        "The target PDF directory could not be found:\n"
        f"{TARGET_DIRECTORY}"
    )

if not TARGET_DIRECTORY.is_dir():
    raise NotADirectoryError(
        "The configured target path is not a directory:\n"
        f"{TARGET_DIRECTORY}"
    )

REPORTS_DIRECTORY.mkdir(
    parents=True,
    exist_ok=True
)

sample_list = get_file_names(
    directory=TARGET_DIRECTORY,
    rejected_directory=REJECTED_DIRECTORY
)
results = []

print("Found files:")

for file_path in sample_list:
    print(file_path)

print("\nProcessing...\n")


for file_path in sample_list:

    try:
        standard = process_file(file_path)

    except Exception as error:
        print(
            f"Processing failed for {file_path}: "
            f"{type(error).__name__}: {error}"
        )

        results.append(
            {
                "Original File": file_path,
                "Source Filename": os.path.basename(file_path),
                "Filename Hint Code": None,
                "Filename Hint Year": None,
                "Filename Hint Amendment": None,
                "Filename Hint Amendment Year": None,
                "Raw Code": None,
                "Normalised Code": None,
                "Extracted Year": None,
                "Amendment": None,
                "Amendment Year": None,
                "Score": 0,
                "Extracted Text Length": 0,
                "Generated Filename": None,
                "Proposed Path": None,
                "Status": "REVIEW_REQUIRED",
                "Reasons": (
                    f"Processing error: "
                    f"{type(error).__name__}: {error}"
                ),
                "Rename Result": "Not renamed",
                "Open File": "",
                "Manual Decision": "",
                "Approved Filename": "",
                "Reviewer Notes": "",
                "Approval Result": "",
                "Approved By": "",
                "Approved At": "",
                "Final Path": ""
            }
        )

        continue

    proposed_path = standard.proposed_path

    print(
        f"{file_path}\n"
        f"  Proposed filename: {standard.filename}\n"
        f"  Status: {standard.status}\n"
        f"  Reasons: {standard.reasons_text() or 'None'}"
    )

    # Default result until rename logic runs.
    rename_result = "Not renamed"

    # --------------------------------------------------------
    # Dry run
    # --------------------------------------------------------

    if DRY_RUN:

        if (
            standard.status == "SUCCESS"
            and proposed_path
        ):
            rename_result = "Dry run - would rename"

            print(
                "  Would rename:\n"
                f"    {file_path}\n"
                "  to:\n"
                f"    {proposed_path}"
            )

        else:
            rename_result = (
                f"Dry run - not eligible because status is "
                f"{standard.status}"
            )

            print(
                "  Would not rename automatically."
            )

    # --------------------------------------------------------
    # Real rename
    # --------------------------------------------------------

    else:

        # Only SUCCESS records are eligible for automatic renaming.
        if (
            standard.status == "SUCCESS"
            and standard.filename
        ):
            success, message = rename_file(
                file_path,
                standard
            )

            rename_result = message

            if success:
                print(f"  Rename result: {message}")

            else:
                standard.status = "REVIEW_REQUIRED"
                standard.add_reason(message)

                print(f"  Rename failed: {message}")

        else:
            rename_result = (
                f"Not renamed because status is "
                f"{standard.status}"
            )

            print(
                "  Not renamed automatically."
            )

    standard.rename_result = rename_result

    # --------------------------------------------------------
    # Add every file to the review report
    # --------------------------------------------------------

    results.append(
        {
            "Original File": file_path,
            "Source Filename": standard.source_filename,
            "Filename Hint Code": standard.filename_hint_code,
            "Filename Hint Year": standard.filename_hint_year,
            "Filename Hint Amendment":
                standard.filename_hint_amendment,
            "Filename Hint Amendment Year":
                standard.filename_hint_amendment_year,
            "Raw Code": standard.raw_code,
            "Normalised Code": standard.normalised_code,
            "Extracted Year": standard.year,
            "Amendment": standard.amendment,
            "Amendment Year": standard.amendment_year,
            "Score": standard.score,
            "Extracted Text Length":
                standard.extracted_text_length,
            "Generated Filename": standard.filename,
            "Proposed Path": proposed_path,
            "Status": standard.status,
            "Reasons": standard.reasons_text(),
            "Rename Result": rename_result,
            "Open File": "",
            "Manual Decision": "",
            "Approved Filename": "",
            "Reviewer Notes": "",
            "Approval Result": "",
            "Approved By": "",
            "Approved At": "",
            "Final Path": ""
        }
    )

    print()


# ------------------------------------------------------------
# Build final report
# ------------------------------------------------------------

final_list = pd.DataFrame(results)

print("\nFinal List:")
print(final_list)

print("\nStatus Summary:")

if not final_list.empty:
    print(
        final_list["Status"].value_counts(
            dropna=False
        )
    )


# ------------------------------------------------------------
# Export results
# ------------------------------------------------------------

if EXPORT_RESULTS and not final_list.empty:

    summary = (
        final_list["Status"]
        .value_counts(dropna=False)
        .rename_axis("Status")
        .reset_index(name="Count")
    )

    status_counts = (
        final_list["Status"]
        .value_counts(dropna=False)
        .to_dict()
    )

    audit_details = (
        f"Processed={len(final_list)}; "
        f"SUCCESS={status_counts.get('SUCCESS', 0)}; "
        f"REVIEW_REQUIRED="
        f"{status_counts.get('REVIEW_REQUIRED', 0)}; "
        f"NO_CANDIDATE="
        f"{status_counts.get('NO_CANDIDATE', 0)}; "
        f"Target={TARGET_DIRECTORY}"
    )

    audit_log = pd.DataFrame(
        [
            {
                "Timestamp": RUN_STARTED_AT
                ,
                "Event": "PROCESSING_RUN",
                "User": CURRENT_USER,
                "Dry Run": DRY_RUN,
                "Script": "Script.py",
                "Details": audit_details
            }
        ]
    )

    def write_results_workbook(output_path):
        """
        Export results and apply Excel formatting, tables,
        filters, and the manual-decision dropdown.
        """

        # ----------------------------------------------------
        # Write the raw worksheet data
        # ----------------------------------------------------

        with pd.ExcelWriter(
            output_path,
            engine="openpyxl"
        ) as writer:

            final_list.to_excel(
                writer,
                sheet_name="All Results",
                index=False
            )

            summary.to_excel(
                writer,
                sheet_name="Summary",
                index=False
            )

            audit_log.to_excel(
                writer,
                sheet_name="Audit Log",
                index=False
            )

        # ----------------------------------------------------
        # Reopen with openpyxl to add Excel features
        # ----------------------------------------------------

        workbook = load_workbook(
            output_path
        )

        results_sheet = workbook[
            "All Results"
        ]

        summary_sheet = workbook[
            "Summary"
        ]

        audit_sheet = workbook[
            "Audit Log"
        ]

        # ----------------------------------------------------
        # Convert All Results into an Excel table
        # ----------------------------------------------------

        if results_sheet.max_row >= 2:

            results_table_reference = (
                f"A1:"
                f"{get_column_letter(results_sheet.max_column)}"
                f"{results_sheet.max_row}"
            )

            results_table = Table(
                displayName="tblRenameResults",
                ref=results_table_reference
            )

            results_table_style = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )

            results_table.tableStyleInfo = (
                results_table_style
            )

            results_sheet.add_table(
                results_table
            )

        # ----------------------------------------------------
        # Convert Summary into an Excel table
        # ----------------------------------------------------

        if summary_sheet.max_row >= 2:

            summary_table_reference = (
                f"A1:"
                f"{get_column_letter(summary_sheet.max_column)}"
                f"{summary_sheet.max_row}"
            )

            summary_table = Table(
                displayName="tblRenameSummary",
                ref=summary_table_reference
            )

            summary_table_style = TableStyleInfo(
                name="TableStyleMedium4",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )

            summary_table.tableStyleInfo = (
                summary_table_style
            )

            summary_sheet.add_table(
                summary_table
            )

        # ----------------------------------------------------
        # Convert Audit Log into an Excel table
        # ----------------------------------------------------

        if audit_sheet.max_row >= 2:

            audit_table_reference = (
                f"A1:"
                f"{get_column_letter(audit_sheet.max_column)}"
                f"{audit_sheet.max_row}"
            )

            audit_table = Table(
                displayName="tblAuditLog",
                ref=audit_table_reference
            )

            audit_table_style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )

            audit_table.tableStyleInfo = (
                audit_table_style
            )

            audit_sheet.add_table(
                audit_table
            )

        # ----------------------------------------------------
        # Freeze the heading row
        # ----------------------------------------------------

        results_sheet.freeze_panes = "A2"
        summary_sheet.freeze_panes = "A2"
        audit_sheet.freeze_panes = "A2"

        # ----------------------------------------------------
        # Locate columns by heading
        # ----------------------------------------------------

        headers = {
            cell.value: cell.column
            for cell in results_sheet[1]
        }

        decision_column = headers.get(
            "Manual Decision"
        )

        approved_filename_column = headers.get(
            "Approved Filename"
        )

        reviewer_notes_column = headers.get(
            "Reviewer Notes"
        )

        status_column = headers.get(
            "Status"
        )

        reasons_column = headers.get(
            "Reasons"
        )

        original_file_column = headers.get(
            "Original File"
        )

        proposed_path_column = headers.get(
            "Proposed Path"
        )

        open_file_column = headers.get(
            "Open File"
        )

        # ----------------------------------------------------
        # Add relative Open File hyperlinks
        # ----------------------------------------------------

        if (
            original_file_column
            and open_file_column
        ):

            workbook_directory = Path(
                output_path
            ).resolve().parent

            for row_number in range(
                2,
                results_sheet.max_row + 1
            ):

                original_file_value = results_sheet.cell(
                    row=row_number,
                    column=original_file_column
                ).value

                proposed_path_value = None

                if proposed_path_column:
                    proposed_path_value = results_sheet.cell(
                        row=row_number,
                        column=proposed_path_column
                    ).value

                original_path = None
                proposed_path = None

                if original_file_value:
                    original_path = Path(
                        str(original_file_value)
                    )

                if proposed_path_value:
                    proposed_path = Path(
                        str(proposed_path_value)
                    )

                if (
                    proposed_path is not None
                    and proposed_path.exists()
                ):
                    file_to_open = proposed_path

                elif (
                    original_path is not None
                    and original_path.exists()
                ):
                    file_to_open = original_path

                else:
                    file_to_open = None

                open_file_cell = results_sheet.cell(
                    row=row_number,
                    column=open_file_column
                )

                if file_to_open is None:
                    open_file_cell.value = "File not found"
                    open_file_cell.hyperlink = None

                    open_file_cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )

                else:
                    try:
                        relative_path = os.path.relpath(
                            file_to_open.resolve(),
                            start=workbook_directory
                        )

                        relative_link = relative_path.replace(
                            "\\",
                            "/"
                        )

                        encoded_relative_link = quote(
                            relative_link,
                            safe="/:."
                        )

                        open_file_cell.value = "Open File"
                        open_file_cell.hyperlink = (
                            encoded_relative_link
                        )
                        open_file_cell.style = "Hyperlink"

                        open_file_cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center"
                        )

                        print(
                            f"Relative hyperlink: "
                            f"{encoded_relative_link}"
                        )

                    except ValueError as error:
                        open_file_cell.value = (
                            "Relative link unavailable"
                        )

                        open_file_cell.hyperlink = None

                        open_file_cell.alignment = Alignment(
                            horizontal="center",
                            vertical="center"
                        )

                        print(
                            f"Could not create hyperlink for "
                            f"{file_to_open}: {error}"
                        )
        # ----------------------------------------------------
        # Add APPROVE / REJECT / HOLD dropdown
        # ----------------------------------------------------

        if decision_column:

            decision_validation = DataValidation(
                type="list",
                formula1='"APPROVE,REJECT,HOLD"',
                allow_blank=True
            )

            decision_validation.error = (
                "Choose APPROVE, REJECT, or HOLD."
            )

            decision_validation.errorTitle = (
                "Invalid manual decision"
            )

            decision_validation.prompt = (
                "Select APPROVE, REJECT, or HOLD."
            )

            decision_validation.promptTitle = (
                "Manual review decision"
            )

            decision_validation.showErrorMessage = True
            decision_validation.showInputMessage = True

            results_sheet.add_data_validation(
                decision_validation
            )

            decision_column_letter = (
                get_column_letter(
                    decision_column
                )
            )

            # Apply the dropdown to all current result rows.
            decision_validation.add(
                f"{decision_column_letter}2:"
                f"{decision_column_letter}"
                f"{results_sheet.max_row}"
            )

        # ----------------------------------------------------
        # Apply visual formatting to manual-review columns
        # ----------------------------------------------------

        manual_review_fill = PatternFill(
            fill_type="solid",
            fgColor="FFF2CC"
        )

        if decision_column:

            for row_number in range(
                2,
                results_sheet.max_row + 1
            ):
                results_sheet.cell(
                    row=row_number,
                    column=decision_column
                ).fill = manual_review_fill

        if approved_filename_column:

            for row_number in range(
                2,
                results_sheet.max_row + 1
            ):
                results_sheet.cell(
                    row=row_number,
                    column=approved_filename_column
                ).fill = manual_review_fill

        if reviewer_notes_column:

            for row_number in range(
                2,
                results_sheet.max_row + 1
            ):
                results_sheet.cell(
                    row=row_number,
                    column=reviewer_notes_column
                ).fill = manual_review_fill

        # ----------------------------------------------------
        # Highlight review and error statuses
        # ----------------------------------------------------

        review_fill = PatternFill(
            fill_type="solid",
            fgColor="FCE4D6"
        )

        no_candidate_fill = PatternFill(
            fill_type="solid",
            fgColor="F4CCCC"
        )

        success_fill = PatternFill(
            fill_type="solid",
            fgColor="E2F0D9"
        )

        if status_column:

            for row_number in range(
                2,
                results_sheet.max_row + 1
            ):

                status_cell = results_sheet.cell(
                    row=row_number,
                    column=status_column
                )

                status_value = str(
                    status_cell.value or ""
                ).strip().upper()

                if status_value == "SUCCESS":
                    status_cell.fill = success_fill

                elif status_value == "REVIEW_REQUIRED":
                    status_cell.fill = review_fill

                elif status_value == "NO_CANDIDATE":
                    status_cell.fill = no_candidate_fill

        # ----------------------------------------------------
        # Wrap long text columns
        # ----------------------------------------------------

        wrap_columns = [
            "Original File",
            "Proposed Path",
            "Reasons",
            "Rename Result",
            "Reviewer Notes",
            "Approval Result",
            "Final Path"
        ]

        for heading in wrap_columns:

            column_number = headers.get(
                heading
            )

            if not column_number:
                continue

            for row_number in range(
                2,
                results_sheet.max_row + 1
            ):
                results_sheet.cell(
                    row=row_number,
                    column=column_number
                ).alignment = Alignment(
                    wrap_text=True,
                    vertical="top"
                )

        # ----------------------------------------------------
        # Set sensible column widths
        # ----------------------------------------------------

        preferred_widths = {
            "Original File": 45,
            "Source Filename": 35,
            "Filename Hint Code": 28,
            "Filename Hint Year": 18,
            "Filename Hint Amendment": 22,
            "Filename Hint Amendment Year": 26,
            "Raw Code": 24,
            "Normalised Code": 28,
            "Extracted Year": 16,
            "Amendment": 14,
            "Amendment Year": 18,
            "Score": 12,
            "Extracted Text Length": 22,
            "Generated Filename": 38,
            "Proposed Path": 50,
            "Status": 20,
            "Reasons": 50,
            "Rename Result": 35,
            "Manual Decision": 20,
            "Approved Filename": 38,
            "Reviewer Notes": 50,
            "Open File": 14,
            "Approval Result": 45,
            "Approved By": 22,
            "Approved At": 22,
            "Final Path": 50
        }

        for heading, width in preferred_widths.items():

            column_number = headers.get(
                heading
            )

            if column_number:

                results_sheet.column_dimensions[
                    get_column_letter(
                        column_number
                    )
                ].width = width

        summary_sheet.column_dimensions["A"].width = 24
        summary_sheet.column_dimensions["B"].width = 12

        # ----------------------------------------------------
        # Improve header appearance
        # ----------------------------------------------------

        for cell in results_sheet[1]:
            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        results_sheet.row_dimensions[1].height = 32

        for cell in summary_sheet[1]:
            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        summary_sheet.row_dimensions[1].height = 24

        # ----------------------------------------------------
        # Format Audit Log
        # ----------------------------------------------------

        audit_widths = {
            "A": 22,
            "B": 22,
            "C": 22,
            "D": 12,
            "E": 24,
            "F": 80
        }

        for column_letter, width in audit_widths.items():
            audit_sheet.column_dimensions[
                column_letter
            ].width = width

        for cell in audit_sheet[1]:
            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        audit_sheet.row_dimensions[1].height = 28

        for row_number in range(
            2,
            audit_sheet.max_row + 1
        ):
            audit_sheet.cell(
                row=row_number,
                column=6
            ).alignment = Alignment(
                wrap_text=True,
                vertical="top"
            )

        # ----------------------------------------------------
        # Save completed workbook
        # ----------------------------------------------------

        workbook.save(
            output_path
        )

    try:

        write_results_workbook(
            RESULTS_WORKBOOK
        )

        print(
            "\nResults exported to:"
        )

        print(
            f"  {RESULTS_WORKBOOK.resolve()}"
        )

    except PermissionError:

        fallback_timestamp = (
            datetime.now().strftime(
                "%y-%m-%d_%H-%M-%S"
            )
        )

        fallback_workbook = (
            REPORTS_DIRECTORY
            / f"RenameResults_{fallback_timestamp}.xlsx"
        )

        write_results_workbook(
            fallback_workbook
        )

        print(
            "\nThe main workbook could not be overwritten."
        )

        print(
            "It may currently be open in Excel."
        )

        print(
            "\nResults were saved instead to:"
        )

        print(
            f"  {fallback_workbook.resolve()}"
        )

elif not EXPORT_RESULTS:

    print(
        "\nEXPORT_RESULTS is disabled. "
        "No workbook was created."
    )

else:

    print(
        "\nThere were no results to export."
    )



# ------------------------------------------------------------
# Optional targeted debugging
# ------------------------------------------------------------

# Uncomment these when you specifically want detailed debugging.

# test_single_file(
#     r"Test Docs\Documents\NR_L2_ELP_27716-01_2023.pdf"
# )

# test_single_file(
#     r"Test Docs\Documents\EN_50238-3_2013.pdf"
# )

# test_single_file(
#     r"Test Docs\Documents\IEC_61000-1-2_2001.pdf"
# )